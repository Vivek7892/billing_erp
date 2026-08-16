from rest_framework import viewsets, status, filters
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import authenticate
from django.http import HttpResponse
from django.db.models import Sum, Count, Q, F, Avg
from django.utils import timezone
from datetime import timedelta, date
from decimal import Decimal
from io import BytesIO
import json

import csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from .models import *
from .serializers import *
from .permissions import IsAdmin, IsAdminOrReadOnly, IsCashierOrAdmin
from .supabase_storage import SupabaseStorageError, upload_shop_logo
from django_filters.rest_framework import DjangoFilterBackend


def _report_value(value):
    if value is None:
        return ''
    if isinstance(value, float):
        return f'{value:.2f}'
    if isinstance(value, Decimal):
        return f'{value:.2f}'
    return str(value)


def _rupee(value):
    """Prefix a numeric string with Rs. (safe for Helvetica fallback font)."""
    s = str(value).strip()
    if not s or s in ('', '0.00', '0'):
        return s
    # Only prefix if it looks like a number
    try:
        float(s.replace(',', ''))
        return f'Rs.{s}'
    except ValueError:
        return s


def _report_response(filename, content, content_type):
    response = HttpResponse(content, content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _export_report_xlsx(title, headers, rows, filename, summary_rows=None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title[:31]
    row_index = 1
    sheet.cell(row=row_index, column=1, value=title)
    sheet.cell(row=row_index, column=1).font = Font(bold=True, size=14)
    row_index += 2
    if summary_rows:
        for label, value in summary_rows:
            sheet.cell(row=row_index, column=1, value=label)
            sheet.cell(row=row_index, column=2, value=_report_value(value))
            sheet.cell(row=row_index, column=1).font = Font(bold=True)
            row_index += 1
        row_index += 1
    for col_index, header in enumerate(headers, 1):
        cell = sheet.cell(row=row_index, column=col_index, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='000000')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
        )
    header_row = row_index
    row_index += 1
    for row in rows:
        for col_index, value in enumerate(row, 1):
            cell = sheet.cell(row=row_index, column=col_index, value=_report_value(value))
            cell.alignment = Alignment(horizontal='left' if col_index == 1 else 'right', vertical='center')
            cell.border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
            )
        row_index += 1
    for col in sheet.columns:
        width = max(len(_report_value(cell.value)) for cell in col if cell.value is not None)
        sheet.column_dimensions[col[0].column_letter].width = min(max(width + 2, 12), 30)
    sheet.freeze_panes = f'A{header_row + 1}'
    buffer = BytesIO()
    workbook.save(buffer)
    return _report_response(filename, buffer.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _export_report_pdf(title, headers, rows, filename, summary_rows=None, request_user=None):
    """
    Generate a clean A4 report PDF with a compact, non-repeating header/footer.

    Header:
      - Shop logo (when available)
      - Shop name
      - Address / phone / email / GSTIN
      - Report title
      - One combined generated date & time line

    Footer:
      - Page number only

    Date/time and contact information are intentionally shown only once in the
    header so the PDF does not repeat the same details unnecessarily.
    """
    from .models import Setting
    import os
    from datetime import datetime

    def get_setting(key, default=''):
        if request_user is not None and getattr(request_user, 'business', None):
            setting = Setting.objects.filter(
                business=request_user.business,
                key=key,
            ).first()
        else:
            setting = Setting.objects.filter(key=key).first()
        return setting.value if setting else default

    shop_name = get_setting('shop_name', 'ShopEase POS')
    shop_address = get_setting('shop_address', '')
    shop_phone = get_setting('shop_phone', '')
    shop_email = get_setting('shop_email', '')
    shop_gstin = get_setting('shop_gstin', '')
    shop_logo_url = get_setting('shop_logo', '')

    now = timezone.localtime(timezone.now())
    generated_at = now.strftime('%d-%m-%Y %I:%M %p')

    # ---------------------------------------------------------
    # COLORS
    # ---------------------------------------------------------
    HEADER_BLUE = colors.HexColor('#1e40af')
    HEADER_BG = colors.HexColor('#eff6ff')
    ROW_ALT = colors.HexColor('#f8fafc')
    BORDER = colors.HexColor('#cbd5e1')
    TEXT = colors.HexColor('#1e293b')
    SUBTLE = colors.HexColor('#64748b')

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=10 * mm,
        bottomMargin=14 * mm,
        title=title,
        author=shop_name,
    )

    story = []
    body_w = A4[0] - 24 * mm

    # ---------------------------------------------------------
    # LOGO
    # ---------------------------------------------------------
    logo_img = None
    try:
        from django.conf import settings as django_settings
        from reportlab.platypus import Image as RLImage
        import glob

        media_root = str(django_settings.MEDIA_ROOT)
        media_url = django_settings.MEDIA_URL

        if shop_logo_url and shop_logo_url.startswith(media_url):
            rel = shop_logo_url[len(media_url):]
            full = os.path.join(media_root, rel)

            if os.path.exists(full):
                logo_img = RLImage(
                    full,
                    width=14 * mm,
                    height=14 * mm,
                )

        if not logo_img:
            matches = glob.glob(
                os.path.join(media_root, 'shop_logos', '*')
            )
            if matches:
                logo_img = RLImage(
                    matches[0],
                    width=14 * mm,
                    height=14 * mm,
                )
    except Exception:
        logo_img = None

    # ---------------------------------------------------------
    # STYLES
    # ---------------------------------------------------------
    name_style = ParagraphStyle(
        'report_shop_name',
        fontName='Helvetica-Bold',
        fontSize=15,
        leading=17,
        alignment=TA_LEFT,
        textColor=HEADER_BLUE,
    )

    sub_style = ParagraphStyle(
        'report_shop_sub',
        fontName='Helvetica',
        fontSize=7.5,
        leading=10,
        alignment=TA_LEFT,
        textColor=SUBTLE,
    )

    title_style = ParagraphStyle(
        'report_title',
        fontName='Helvetica-Bold',
        fontSize=13,
        leading=15,
        alignment=TA_RIGHT,
        textColor=TEXT,
    )

    meta_style = ParagraphStyle(
        'report_meta',
        fontName='Helvetica',
        fontSize=7.5,
        leading=10,
        alignment=TA_RIGHT,
        textColor=SUBTLE,
    )

    # ---------------------------------------------------------
    # HEADER
    # ---------------------------------------------------------
    shop_lines = [
        Paragraph(shop_name or 'ShopEase POS', name_style)
    ]

    if shop_address:
        shop_lines.append(
            Paragraph(
                shop_address.replace('\n', ', '),
                sub_style,
            )
        )

    contact_parts = []
    if shop_phone:
        contact_parts.append(f'Phone: {shop_phone}')
    if shop_email:
        contact_parts.append(f'Email: {shop_email}')
    if shop_gstin:
        contact_parts.append(f'GSTIN: {shop_gstin}')

    if contact_parts:
        shop_lines.append(
            Paragraph(
                '  |  '.join(contact_parts),
                sub_style,
            )
        )

    # Date and time appear ONCE, as a single compact line.
    right_lines = [
        Paragraph(title, title_style),
        Spacer(1, 1.5 * mm),
        Paragraph(f'Generated: {generated_at}', meta_style),
    ]

    if logo_img:
        logo_col_w = 16 * mm
        info_col_w = body_w * 0.57
        right_col_w = body_w - logo_col_w - info_col_w

        header_tbl = Table(
            [[logo_img, shop_lines, right_lines]],
            colWidths=[
                logo_col_w,
                info_col_w,
                right_col_w,
            ],
        )

        header_tbl.setStyle(
            TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
                ('LEFTPADDING', (1, 0), (1, 0), 5),
            ])
        )
    else:
        header_tbl = Table(
            [[shop_lines, right_lines]],
            colWidths=[
                body_w * 0.60,
                body_w * 0.40,
            ],
        )

        header_tbl.setStyle(
            TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ])
        )

    story.append(header_tbl)
    story.append(Spacer(1, 2.5 * mm))

    rule_tbl = Table([['']], colWidths=[body_w])
    rule_tbl.setStyle(
        TableStyle([
            ('LINEBELOW', (0, 0), (-1, -1), 1.0, HEADER_BLUE),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ])
    )

    story.append(rule_tbl)
    story.append(Spacer(1, 4 * mm))

    # ---------------------------------------------------------
    # SUMMARY
    # ---------------------------------------------------------
    if summary_rows:
        s_style = ParagraphStyle(
            'report_summary_label',
            fontName='Helvetica',
            fontSize=8.5,
            leading=11,
            textColor=TEXT,
        )

        s_bold = ParagraphStyle(
            'report_summary_value',
            fontName='Helvetica-Bold',
            fontSize=8.5,
            leading=11,
            alignment=TA_RIGHT,
            textColor=TEXT,
        )

        summary_data = [
            [
                Paragraph(str(label), s_style),
                Paragraph(_rupee(_report_value(value)), s_bold),
            ]
            for label, value in summary_rows
        ]

        summary_tbl = Table(
            summary_data,
            colWidths=[80 * mm, 60 * mm],
        )

        summary_tbl.setStyle(
            TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), HEADER_BG),
                ('BOX', (0, 0), (-1, -1), 0.5, BORDER),
                ('LINEBELOW', (0, 0), (-1, -2), 0.3, BORDER),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ])
        )

        story.append(summary_tbl)
        story.append(Spacer(1, 5 * mm))

    # ---------------------------------------------------------
    # DATA TABLE
    # ---------------------------------------------------------
    h_style = ParagraphStyle(
        'report_table_header',
        fontName='Helvetica-Bold',
        fontSize=7.5,
        leading=9,
        alignment=TA_CENTER,
        textColor=colors.white,
    )

    c_style = ParagraphStyle(
        'report_table_cell',
        fontName='Helvetica',
        fontSize=7.5,
        leading=9,
        textColor=TEXT,
    )

    c_right = ParagraphStyle(
        'report_table_cell_right',
        fontName='Helvetica',
        fontSize=7.5,
        leading=9,
        alignment=TA_RIGHT,
        textColor=TEXT,
    )

    header_row = [
        Paragraph(str(h), h_style)
        for h in headers
    ]

    # Currency is applied only to columns whose header indicates an amount.
    currency_keywords = (
        'amount', 'revenue', 'cost', 'profit', 'total',
        'sales', 'discount', 'tax', 'gst', 'outstanding',
        'credit', 'expense', 'payment', 'price', 'value',
    )

    data_rows = []
    for row in rows:
        formatted_row = []

        for j, value in enumerate(row):
            header_name = str(headers[j]).lower() if j < len(headers) else ''
            is_currency = any(
                keyword in header_name
                for keyword in currency_keywords
            )

            display_value = _rupee(_report_value(value)) if is_currency else _report_value(value)

            formatted_row.append(
                Paragraph(
                    display_value,
                    c_right if j > 0 else c_style,
                )
            )

        data_rows.append(formatted_row)

    if headers:
        col_width = body_w / len(headers)
        table = Table(
            [header_row] + data_rows,
            colWidths=[col_width] * len(headers),
            repeatRows=1,
            splitByRow=1,
        )

        tbl_style = [
            ('BACKGROUND', (0, 0), (-1, 0), HEADER_BLUE),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7.5),
            ('BOX', (0, 0), (-1, -1), 0.5, BORDER),
            ('LINEBELOW', (0, 0), (-1, 0), 0.5, BORDER),
            ('LINEBELOW', (0, 1), (-1, -1), 0.3, BORDER),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]

        for i in range(1, len(data_rows) + 1):
            if i % 2 == 0:
                tbl_style.append(
                    (
                        'BACKGROUND',
                        (0, i),
                        (-1, i),
                        ROW_ALT,
                    )
                )

        table.setStyle(TableStyle(tbl_style))
        story.append(table)
    else:
        story.append(
            Paragraph('No report data available.', c_style)
        )

    # ---------------------------------------------------------
    # PAGE FOOTER
    # ---------------------------------------------------------
    def draw_page_footer(canvas, doc):
        canvas.saveState()

        page_width, _ = A4

        canvas.setStrokeColor(BORDER)
        canvas.setLineWidth(0.5)
        canvas.line(
            12 * mm,
            8 * mm,
            page_width - 12 * mm,
            8 * mm,
        )

        footer_style = ParagraphStyle(
            'page_footer',
            fontName='Helvetica',
            fontSize=7,
            leading=8,
            alignment=TA_CENTER,
            textColor=SUBTLE,
        )

        # Footer contains shop name, address and contact details.
        # Date/time and report title are intentionally not repeated here.
        contact_parts = []

        if shop_name:
            contact_parts.append(shop_name)

        if shop_address:
            contact_parts.append(shop_address.replace('\n', ', '))

        if shop_phone:
            contact_parts.append(f'Phone: {shop_phone}')

        if shop_email:
            contact_parts.append(f'Email: {shop_email}')

        contact_text = '  |  '.join(contact_parts)

        footer_contact = Paragraph(
            contact_text,
            footer_style,
        )

        footer_contact.wrapOn(canvas, body_w, 8 * mm)
        footer_contact.drawOn(
            canvas,
            12 * mm,
            5 * mm,
        )

        # Page number only below the contact line.
        page_text = Paragraph(
            f'Page {doc.page}',
            footer_style,
        )

        page_text.wrapOn(canvas, body_w, 4 * mm)
        page_text.drawOn(
            canvas,
            12 * mm,
            2 * mm,
        )

        canvas.restoreState()

    # ---------------------------------------------------------
    # BUILD PDF
    # ---------------------------------------------------------
    doc.build(
        story,
        onFirstPage=draw_page_footer,
        onLaterPages=draw_page_footer,
    )

    buffer.seek(0)

    return _report_response(
        filename,
        buffer.getvalue(),
        'application/pdf',
    )


class LoginView(APIView):
    permission_classes = []

    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')
        user = authenticate(username=username, password=password)
        if not user or not user.is_active:
            return Response({'error': 'Invalid credentials'}, status=400)
        refresh = RefreshToken.for_user(user)
        return Response({
            'access': str(refresh.access_token),
            'refresh': str(refresh),
            'user': UserSerializer(user).data,
        })


class SuperAdminLoginView(APIView):
    permission_classes = []

    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')
        user = authenticate(username=username, password=password)
        if not user or not user.is_active or not user.is_superuser:
            return Response({'error': 'Invalid super admin credentials'}, status=400)
        refresh = RefreshToken.for_user(user)
        return Response({
            'access': str(refresh.access_token),
            'refresh': str(refresh),
            'user': {'id': user.id, 'username': user.username, 'is_superuser': True},
        })


class SuperAdminDashboardView(APIView):
    def get(self, request):
        if not request.user.is_superuser:
            return Response({'error': 'Forbidden'}, status=403)
        businesses = Business.objects.annotate(
            user_count=Count('members', distinct=True),
            invoice_count=Count('invoice', distinct=True),
            total_revenue=Sum('invoice__grand_total'),
        ).values(
            'id', 'name', 'owner_name', 'mobile', 'email',
            'business_type', 'created_at', 'user_count', 'invoice_count', 'total_revenue'
        ).order_by('-created_at')
        return Response({
            'total_businesses': Business.objects.count(),
            'total_users': User.objects.count(),
            'total_invoices': Invoice.objects.count(),
            'total_revenue': Invoice.objects.aggregate(t=Sum('grand_total'))['t'] or 0,
            'businesses': list(businesses),
        })


class MeView(APIView):
    def get(self, request):
        return Response(UserSerializer(request.user).data)

    def put(self, request):
        serializer = UserSerializer(request.user, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


class ChangePasswordView(APIView):
    def post(self, request):
        current = request.data.get('current_password', '')
        new_pwd = request.data.get('new_password', '')
        if not request.user.check_password(current):
            return Response({'detail': 'Current password is incorrect.'}, status=400)
        if len(new_pwd) < 6:
            return Response({'detail': 'Password must be at least 6 characters.'}, status=400)
        request.user.set_password(new_pwd)
        request.user.save()
        return Response({'status': 'changed'})


class UserViewSet(viewsets.ModelViewSet):
    serializer_class = UserSerializer
    permission_classes = [IsAdmin]
    filter_backends = [filters.SearchFilter]
    search_fields = ['username', 'email', 'first_name', 'last_name']

    def get_queryset(self):
        return User.objects.filter(business=self.request.user.business)

    def perform_create(self, serializer):
        serializer.save(business=self.request.user.business)


class CategoryViewSet(viewsets.ModelViewSet):
    serializer_class = CategorySerializer
    permission_classes = [IsAdminOrReadOnly]
    filter_backends = [filters.SearchFilter]
    search_fields = ['name']

    def get_queryset(self):
        return Category.objects.filter(business=self.request.user.business)

    def perform_create(self, serializer):
        serializer.save(business=self.request.user.business)


class SupplierViewSet(viewsets.ModelViewSet):
    serializer_class = SupplierSerializer
    permission_classes = [IsAdminOrReadOnly]
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'phone', 'email']

    def get_queryset(self):
        return Supplier.objects.filter(business=self.request.user.business)

    def perform_create(self, serializer):
        serializer.save(business=self.request.user.business)


class ProductViewSet(viewsets.ModelViewSet):
    serializer_class = ProductSerializer
    permission_classes = [IsAdminOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category', 'status', 'supplier']
    search_fields = ['name', 'sku', 'barcode', 'brand']
    ordering_fields = ['name', 'selling_price', 'current_stock', 'created_at']

    def get_queryset(self):
        return Product.objects.select_related('category', 'supplier').filter(
            business=self.request.user.business
        )

    def perform_create(self, serializer):
        serializer.save(business=self.request.user.business)

    def destroy(self, request, *args, **kwargs):
        if request.user.role != 'admin':
            return Response({'error': 'Permission denied'}, status=403)
        return super().destroy(request, *args, **kwargs)


class CustomerViewSet(viewsets.ModelViewSet):
    serializer_class = CustomerSerializer
    permission_classes = [IsCashierOrAdmin]
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'mobile', 'email']

    def get_queryset(self):
        return Customer.objects.filter(business=self.request.user.business)

    def perform_create(self, serializer):
        serializer.save(business=self.request.user.business)

    @action(detail=True, methods=['get'])
    def bills(self, request, pk=None):
        customer = self.get_object()
        invoices = Invoice.objects.filter(customer=customer).order_by('-created_at')[:20]
        return Response(InvoiceSerializer(invoices, many=True).data)

    @action(detail=True, methods=['post'], url_path='send-reminder')
    def send_reminder(self, request, pk=None):
        customer = self.get_object()
        channel = request.data.get('channel', 'sms')  # sms | whatsapp
        if not customer.mobile:
            return Response({'error': 'Customer has no mobile number'}, status=400)

        from decouple import config as env
        sid = env('TWILIO_ACCOUNT_SID', default='')
        token = env('TWILIO_AUTH_TOKEN', default='')
        if not sid or not token:
            return Response({'error': 'Twilio credentials not configured in .env'}, status=503)

        shop_name = (Setting.objects.filter(key='shop_name').first() or type('', (), {'value': 'ShopEase POS'})()).value
        amount = float(customer.outstanding_amount)
        body = (
            f"Dear {customer.name}, you have an outstanding balance of "
            f"\u20b9{amount:,.2f} at {shop_name}. "
            f"Please clear your dues at your earliest convenience. Thank you!"
        )

        to_number = customer.mobile if customer.mobile.startswith('+') else f'+91{customer.mobile}'
        try:
            from twilio.rest import Client
            client = Client(sid, token)
            if channel == 'whatsapp':
                from_num = env('TWILIO_WHATSAPP_FROM', default='whatsapp:+14155238886')
                msg = client.messages.create(body=body, from_=from_num, to=f'whatsapp:{to_number}')
            else:
                from_num = env('TWILIO_FROM_NUMBER', default='')
                msg = client.messages.create(body=body, from_=from_num, to=to_number)
            return Response({'status': 'sent', 'sid': msg.sid, 'channel': channel})
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class PurchaseViewSet(viewsets.ModelViewSet):
    serializer_class = PurchaseSerializer
    permission_classes = [IsAdminOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['supplier', 'payment_status']
    search_fields = ['invoice_number']

    def get_queryset(self):
        return Purchase.objects.select_related('supplier').prefetch_related('items').filter(
            business=self.request.user.business
        )

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, business=self.request.user.business)


class InvoiceViewSet(viewsets.ModelViewSet):
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['payment_status', 'status', 'payment_method']
    search_fields = ['invoice_number', 'customer_name', 'customer__name']
    ordering_fields = ['created_at', 'grand_total']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.action == 'create':
            return InvoiceCreateSerializer
        return InvoiceSerializer

    def get_permissions(self):
        if self.action in ['destroy']:
            return [IsAdmin()]
        return [IsCashierOrAdmin()]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, business=self.request.user.business)

    def get_queryset(self):
        qs = Invoice.objects.select_related('customer', 'created_by').prefetch_related(
            'items', 'payments'
        ).filter(business=self.request.user.business)
        user = self.request.user
        if user.role == 'cashier':
            qs = qs.filter(created_by=user)
        date_filter = self.request.query_params.get('date_filter')
        today = timezone.now().date()
        if date_filter == 'today':
            qs = qs.filter(created_at__date=today)
        elif date_filter == 'yesterday':
            qs = qs.filter(created_at__date=today - timedelta(days=1))
        elif date_filter == 'this_week':
            qs = qs.filter(created_at__date__gte=today - timedelta(days=7))
        elif date_filter == 'this_month':
            qs = qs.filter(created_at__year=today.year, created_at__month=today.month)
        start = self.request.query_params.get('start_date')
        end = self.request.query_params.get('end_date')
        if start:
            qs = qs.filter(created_at__date__gte=start)
        if end:
            qs = qs.filter(created_at__date__lte=end)
        return qs


class InventoryViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = InventoryTransactionSerializer
    permission_classes = [IsCashierOrAdmin]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['product', 'transaction_type']
    search_fields = ['product__name', 'reference']

    def get_queryset(self):
        return InventoryTransaction.objects.select_related('product', 'created_by').filter(
            business=self.request.user.business
        ).order_by('-created_at')


class SettingViewSet(viewsets.ModelViewSet):
    queryset = Setting.objects.all()
    serializer_class = SettingSerializer
    permission_classes = [IsAdmin]

    def get_queryset(self):
        return Setting.objects.filter(business=self.request.user.business)

    @action(detail=False, methods=['post'])
    def bulk_update(self, request):
        for key, value in request.data.items():
            Setting.objects.update_or_create(business=request.user.business, key=key, defaults={'value': str(value)})
        return Response({'status': 'updated'})

    @action(detail=False, methods=['post'], url_path='upload-logo')
    def upload_logo(self, request):
        """Store the shop logo in Supabase Storage and return its public URL."""
        logo = request.FILES.get('logo')
        if not logo:
            return Response({'detail': 'Choose a logo image first.'}, status=status.HTTP_400_BAD_REQUEST)
        if logo.size > 2 * 1024 * 1024:
            return Response({'detail': 'Logo must be smaller than 2 MB.'}, status=status.HTTP_400_BAD_REQUEST)
        if logo.content_type not in ('image/png', 'image/jpeg', 'image/webp'):
            return Response({'detail': 'Use a PNG, JPG, or WEBP logo image.'}, status=status.HTTP_400_BAD_REQUEST)
        biz_id = str(request.user.business_id or 'default')
        try:
            url = upload_shop_logo(logo, biz_id)
        except SupabaseStorageError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

        Setting.objects.update_or_create(business=request.user.business, key='shop_logo', defaults={'value': url})
        return Response({'url': url})

    @action(detail=False, methods=['get'], permission_classes=[IsCashierOrAdmin])
    def all(self, request):
        defaults = {
            # Business
            'shop_name': '', 'shop_address': '', 'shop_phone': '', 'shop_email': '',
            'shop_gstin': '', 'shop_pan': '', 'shop_logo': '', 'shop_state': '',
            'business_type': 'retail_wholesale', 'cin': '', 'fssai_licence': '',
            'fy_start': 'april',
            # Invoice
            'invoice_prefix': 'INV', 'invoice_start_number': '1001',
            'invoice_template': 'gst_a4', 'invoice_due_days': '15',
            'invoice_terms': '', 'invoice_footer': '',
            'show_discount_col': 'true', 'show_hsn_col': 'true',
            'show_batch_col': 'false', 'show_expiry_col': 'false',
            # GST
            'gst_reg_type': 'regular', 'default_gst_rate': '18',
            'place_of_supply': '', 'tax_on_price': 'exclusive',
            'einvoice_enabled': 'false', 'hsn_summary_on_invoice': 'true',
            'reverse_charge': 'false', 'cess_enabled': 'false',
            # Payment
            'default_payment_method': 'cash', 'round_off': 'nearest',
            'credit_limit_alert': '', 'shop_upi_id': '',
            'shop_bank_details': '',
            # Legacy bank fields (kept for pdf_utils fallback)
            'shop_bank_name': '', 'shop_bank_branch': '',
            'shop_bank_account': '', 'shop_bank_ifsc': '',
            'show_upi_qr_on_invoice': 'true', 'show_upi_qr_on_thermal': 'true',
            'upi_qr_enabled': 'true', 'advance_payment_enabled': 'false',
            # Printer
            'printer_type': 'a4', 'default_printer': '', 'copies_per_bill': '1',
            'auto_print': 'no', 'receipt_footer': 'Thank you for shopping with us!',
            'open_cash_drawer': 'false', 'print_duplicate': 'false',
            # System
            'currency': '₹', 'allow_negative_stock': 'false',
        }
        saved = {s.key: s.value for s in self.get_queryset()}
        return Response({**defaults, **saved})


class DashboardView(APIView):
    def get(self, request):
        biz = request.user.business
        today = timezone.now().date()
        yesterday = today - timedelta(days=1)
        month_start = today.replace(day=1)
        last_month_start = (month_start - timedelta(days=1)).replace(day=1)
        last_month_end = month_start - timedelta(days=1)

        today_invoices = Invoice.objects.filter(business=biz, created_at__date=today, status='completed')
        today_sales = today_invoices.aggregate(total=Sum('grand_total'))['total'] or 0
        today_bills = today_invoices.count()
        today_discount = today_invoices.aggregate(total=Sum('discount_amount'))['total'] or 0
        today_tax = today_invoices.aggregate(total=Sum('tax_amount'))['total'] or 0

        yest_invoices = Invoice.objects.filter(business=biz, created_at__date=yesterday, status='completed')
        yesterday_sales = yest_invoices.aggregate(total=Sum('grand_total'))['total'] or 0
        yesterday_bills = yest_invoices.count()

        month_invoices = Invoice.objects.filter(business=biz, created_at__date__gte=month_start, status='completed')
        month_sales = month_invoices.aggregate(total=Sum('grand_total'))['total'] or 0
        month_bills = month_invoices.count()
        month_discount = month_invoices.aggregate(total=Sum('discount_amount'))['total'] or 0
        month_tax = month_invoices.aggregate(total=Sum('tax_amount'))['total'] or 0

        last_month_sales = Invoice.objects.filter(
            business=biz,
            created_at__date__gte=last_month_start,
            created_at__date__lte=last_month_end,
            status='completed'
        ).aggregate(total=Sum('grand_total'))['total'] or 0

        def calc_profit(items_qs):
            items = items_qs.select_related('product')
            return sum(
                (float(i.unit_price) - float(i.product.purchase_price if i.product else 0)) * float(i.quantity)
                for i in items
            )

        today_profit = calc_profit(InvoiceItem.objects.filter(
            invoice__business=biz, invoice__created_at__date=today, invoice__status='completed'))
        month_profit = calc_profit(InvoiceItem.objects.filter(
            invoice__business=biz, invoice__created_at__date__gte=month_start, invoice__status='completed'))
        yesterday_profit = calc_profit(InvoiceItem.objects.filter(
            invoice__business=biz, invoice__created_at__date=yesterday, invoice__status='completed'))

        total_products = Product.objects.filter(business=biz, status='active').count()
        out_of_stock = Product.objects.filter(business=biz, current_stock__lte=0, status='active').count()
        low_stock_count = Product.objects.filter(
            business=biz, current_stock__gt=0, current_stock__lte=F('minimum_stock'), status='active').count()
        total_customers = Customer.objects.filter(business=biz).count()
        new_customers_today = Customer.objects.filter(business=biz, created_at__date=today).count()
        pending_credit = Customer.objects.filter(business=biz).aggregate(total=Sum('outstanding_amount'))['total'] or 0
        total_suppliers = Supplier.objects.filter(business=biz).count()
        pending_purchases = Purchase.objects.filter(business=biz, payment_status='pending').aggregate(
            total=Sum('total_amount'))['total'] or 0

        sales_7days = []
        for i in range(6, -1, -1):
            d = today - timedelta(days=i)
            day_inv = Invoice.objects.filter(business=biz, created_at__date=d, status='completed')
            day_sales = day_inv.aggregate(total=Sum('grand_total'))['total'] or 0
            day_profit = calc_profit(InvoiceItem.objects.filter(
                invoice__business=biz, invoice__created_at__date=d, invoice__status='completed'))
            sales_7days.append({'date': str(d), 'sales': float(day_sales), 'profit': float(day_profit)})

        monthly_sales = []
        for i in range(5, -1, -1):
            m = (month_start - timedelta(days=i * 30)).replace(day=1)
            total = Invoice.objects.filter(
                business=biz, created_at__year=m.year, created_at__month=m.month, status='completed'
            ).aggregate(total=Sum('grand_total'))['total'] or 0
            bills = Invoice.objects.filter(
                business=biz, created_at__year=m.year, created_at__month=m.month, status='completed'
            ).count()
            monthly_sales.append({'month': m.strftime('%b %Y'), 'total': float(total), 'bills': bills})

        top_products = list(InvoiceItem.objects.filter(
            invoice__business=biz, invoice__status='completed'
        ).values('product_name').annotate(
            total_qty=Sum('quantity'), total_revenue=Sum('total')
        ).order_by('-total_revenue')[:8])

        category_sales = list(InvoiceItem.objects.filter(
            invoice__business=biz,
            invoice__created_at__date__gte=month_start,
            invoice__status='completed',
            product__category__isnull=False
        ).values('product__category__name').annotate(
            total_revenue=Sum('total'), total_qty=Sum('quantity')
        ).order_by('-total_revenue')[:6])

        payment_dist = list(Payment.objects.filter(
            invoice__business=biz, invoice__status='completed'
        ).values('method').annotate(total=Sum('amount'), count=Count('id')))

        hourly_sales = []
        for h in range(0, 24, 2):
            total = Invoice.objects.filter(
                business=biz,
                created_at__date=today,
                created_at__hour__gte=h,
                created_at__hour__lt=h + 2,
                status='completed'
            ).aggregate(total=Sum('grand_total'))['total'] or 0
            hourly_sales.append({'hour': f'{h:02d}:00', 'total': float(total)})

        low_stock_products = list(Product.objects.filter(
            business=biz, current_stock__lte=F('minimum_stock'), status='active'
        ).values('id', 'name', 'current_stock', 'minimum_stock', 'sku').order_by('current_stock')[:10])

        recent_bills = Invoice.objects.filter(business=biz, status='completed').order_by('-created_at')[:10]

        avg_bill_today = float(today_sales) / today_bills if today_bills else 0
        avg_bill_month = float(month_sales) / month_bills if month_bills else 0
        sales_growth = round(((float(month_sales) - float(last_month_sales)) / float(last_month_sales) * 100), 1) if last_month_sales else 0
        profit_margin = round((today_profit / float(today_sales) * 100), 1) if today_sales else 0

        return Response({
            'today_sales': float(today_sales),
            'today_bills': today_bills,
            'today_profit': float(today_profit),
            'today_discount': float(today_discount),
            'today_tax': float(today_tax),
            'yesterday_sales': float(yesterday_sales),
            'yesterday_bills': yesterday_bills,
            'yesterday_profit': float(yesterday_profit),
            'month_sales': float(month_sales),
            'month_bills': month_bills,
            'month_profit': float(month_profit),
            'month_discount': float(month_discount),
            'month_tax': float(month_tax),
            'last_month_sales': float(last_month_sales),
            'total_products': total_products,
            'out_of_stock': out_of_stock,
            'low_stock_count': low_stock_count,
            'total_suppliers': total_suppliers,
            'pending_purchases': float(pending_purchases),
            'total_customers': total_customers,
            'new_customers_today': new_customers_today,
            'pending_credit': float(pending_credit),
            'avg_bill_today': float(avg_bill_today),
            'avg_bill_month': float(avg_bill_month),
            'sales_growth': float(sales_growth),
            'profit_margin': float(profit_margin),
            'sales_7days': sales_7days,
            'monthly_sales': monthly_sales,
            'hourly_sales': hourly_sales,
            'top_products': top_products,
            'category_sales': category_sales,
            'payment_distribution': payment_dist,
            'low_stock_products': low_stock_products,
            'recent_bills': InvoiceSerializer(recent_bills, many=True).data,
        })


class SalesReportView(APIView):
    permission_classes = [IsAdmin]

    def get(self, request):
        start = request.query_params.get('start_date', str(timezone.now().date().replace(day=1)))
        end = request.query_params.get('end_date', str(timezone.now().date()))
        invoices = Invoice.objects.filter(
            created_at__date__gte=start,
            created_at__date__lte=end,
            status='completed'
        )
        summary = invoices.aggregate(
            total_sales=Sum('grand_total'),
            total_discount=Sum('discount_amount'),
            total_tax=Sum('tax_amount'),
            count=Count('id'),
        )
        daily = invoices.values('created_at__date').annotate(
            total=Sum('grand_total'), count=Count('id')
        ).order_by('created_at__date')
        report_format = request.query_params.get('export')
        if report_format in ['pdf', 'xlsx']:
            headers = ['Date', 'Invoices', 'Total Sales']
            rows = [[row['created_at__date'], row['count'], row['total']] for row in daily]
            summary_rows = [
                ('Total Sales', summary['total_sales'] or 0),
                ('Total Discount', summary['total_discount'] or 0),
                ('Total Tax', summary['total_tax'] or 0),
                ('Invoice Count', summary['count'] or 0),
            ]
            filename = f'sales-report.{report_format}'
            if report_format == 'pdf':
                return _export_report_pdf('Sales Report', headers, rows, filename, summary_rows=summary_rows, request_user=request.user)
            return _export_report_xlsx('Sales Report', headers, rows, filename, summary_rows=summary_rows)
        return Response({'summary': summary, 'daily': list(daily)})


class ProductReportView(APIView):
    permission_classes = [IsAdmin]

    def get(self, request):
        start = request.query_params.get('start_date', str(timezone.now().date().replace(day=1)))
        end = request.query_params.get('end_date', str(timezone.now().date()))
        data = InvoiceItem.objects.filter(
            invoice__created_at__date__gte=start,
            invoice__created_at__date__lte=end,
            invoice__status='completed'
        ).values('product_name', 'sku').annotate(
            total_qty=Sum('quantity'),
            total_revenue=Sum('total'),
        ).order_by('-total_qty')
        report_format = request.query_params.get('export')
        if report_format in ['pdf', 'xlsx']:
            headers = ['Product', 'SKU', 'Qty Sold', 'Revenue']
            rows = [[row['product_name'], row['sku'], row['total_qty'], row['total_revenue']] for row in data]
            filename = f'product-report.{report_format}'
            if report_format == 'pdf':
                return _export_report_pdf('Product Sales Report', headers, rows, filename, request_user=request.user)
            return _export_report_xlsx('Product Sales Report', headers, rows, filename)
        return Response(list(data))


class ProfitReportView(APIView):
    permission_classes = [IsAdmin]

    def get(self, request):
        start = request.query_params.get('start_date', str(timezone.now().date().replace(day=1)))
        end = request.query_params.get('end_date', str(timezone.now().date()))
        items = InvoiceItem.objects.filter(
            invoice__created_at__date__gte=start,
            invoice__created_at__date__lte=end,
            invoice__status='completed'
        ).select_related('product')
        result = []
        for item in items:
            cost = float(item.product.purchase_price if item.product else 0) * float(item.quantity)
            revenue = float(item.total)
            result.append({
                'product': item.product_name,
                'qty': float(item.quantity),
                'revenue': revenue,
                'cost': cost,
                'profit': revenue - cost,
            })
        total_revenue = sum(r['revenue'] for r in result)
        total_cost = sum(r['cost'] for r in result)
        report_format = request.query_params.get('export')
        if report_format in ['pdf', 'xlsx']:
            headers = ['Product', 'Qty', 'Revenue', 'Cost', 'Profit']
            rows = [[row['product'], row['qty'], row['revenue'], row['cost'], row['profit']] for row in result]
            summary_rows = [
                ('Total Revenue', total_revenue),
                ('Total Cost', total_cost),
                ('Total Profit', total_revenue - total_cost),
            ]
            filename = f'profit-report.{report_format}'
            if report_format == 'pdf':
                return _export_report_pdf('Profit Report', headers, rows, filename, summary_rows=summary_rows, request_user=request.user)
            return _export_report_xlsx('Profit Report', headers, rows, filename, summary_rows=summary_rows)
        return Response({
            'items': result,
            'total_revenue': total_revenue,
            'total_cost': total_cost,
            'total_profit': total_revenue - total_cost,
        })


class GSTReportView(APIView):
    permission_classes = [IsAdmin]

    def get(self, request):
        start = request.query_params.get('start_date', str(timezone.now().date().replace(day=1)))
        end = request.query_params.get('end_date', str(timezone.now().date()))
        data = InvoiceItem.objects.filter(
            invoice__created_at__date__gte=start,
            invoice__created_at__date__lte=end,
            invoice__status='completed'
        ).values('gst_percent').annotate(
            taxable_amount=Sum(F('total') - F('gst_amount')),
            gst_collected=Sum('gst_amount'),
        ).order_by('gst_percent')
        total_gst = Invoice.objects.filter(
            created_at__date__gte=start,
            created_at__date__lte=end,
            status='completed'
        ).aggregate(total=Sum('tax_amount'))['total'] or 0
        report_format = request.query_params.get('export')
        if report_format in ['pdf', 'xlsx']:
            headers = ['GST %', 'Taxable Amount', 'GST Collected']
            rows = [[row['gst_percent'], row['taxable_amount'], row['gst_collected']] for row in data]
            summary_rows = [('Total GST', total_gst)]
            filename = f'gst-report.{report_format}'
            if report_format == 'pdf':
                return _export_report_pdf('GST Report', headers, rows, filename, summary_rows=summary_rows, request_user=request.user)
            return _export_report_xlsx('GST Report', headers, rows, filename, summary_rows=summary_rows)
        return Response({'by_rate': list(data), 'total_gst': float(total_gst)})


class CustomerCreditReportView(APIView):
    permission_classes = [IsAdmin]

    def get(self, request):
        customers = Customer.objects.filter(outstanding_amount__gt=0).values(
            'id', 'name', 'mobile', 'outstanding_amount', 'credit_limit'
        )
        total = Customer.objects.aggregate(total=Sum('outstanding_amount'))['total'] or 0
        report_format = request.query_params.get('export')
        if report_format in ['pdf', 'xlsx']:
            headers = ['Customer', 'Mobile', 'Outstanding', 'Credit Limit']
            rows = [[row['name'], row['mobile'], row['outstanding_amount'], row['credit_limit']] for row in customers]
            summary_rows = [('Total Outstanding', total)]
            filename = f'customer-credit-report.{report_format}'
            if report_format == 'pdf':
                return _export_report_pdf('Customer Credit Report', headers, rows, filename, summary_rows=summary_rows, request_user=request.user)
            return _export_report_xlsx('Customer Credit Report', headers, rows, filename, summary_rows=summary_rows)
        return Response({'customers': list(customers), 'total_outstanding': float(total)})


class PaymentReportView(APIView):
    permission_classes = [IsAdmin]

    def get(self, request):
        start = request.query_params.get('start_date', str(timezone.now().date().replace(day=1)))
        end = request.query_params.get('end_date', str(timezone.now().date()))
        data = Payment.objects.filter(
            invoice__created_at__date__gte=start,
            invoice__created_at__date__lte=end,
            invoice__status='completed',
            invoice__payment_status__in=['paid', 'partial']
        ).values('method').annotate(total=Sum('amount'), count=Count('id'))
        report_format = request.query_params.get('export')
        if report_format in ['pdf', 'xlsx']:
            headers = ['Method', 'Count', 'Total']
            rows = [[row['method'], row['count'], row['total']] for row in data]
            filename = f'payment-report.{report_format}'
            if report_format == 'pdf':
                return _export_report_pdf('Payment Report', headers, rows, filename, request_user=request.user)
            return _export_report_xlsx('Payment Report', headers, rows, filename)
        return Response(list(data))


class ExpenseCategoryViewSet(viewsets.ModelViewSet):
    serializer_class = ExpenseCategorySerializer
    permission_classes = [IsCashierOrAdmin]
    filter_backends = [filters.SearchFilter]
    search_fields = ['name']

    def get_queryset(self):
        return ExpenseCategory.objects.filter(business=self.request.user.business)

    def perform_create(self, serializer):
        serializer.save(business=self.request.user.business)


class ExpenseViewSet(viewsets.ModelViewSet):
    serializer_class = ExpenseSerializer
    permission_classes = [IsCashierOrAdmin]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['payment_method', 'category']
    search_fields = ['description', 'notes']
    ordering_fields = ['expense_date', 'amount', 'created_at']
    ordering = ['-expense_date']

    def get_queryset(self):
        return Expense.objects.select_related('category', 'created_by').filter(
            business=self.request.user.business
        )

    def perform_create(self, serializer):
        serializer.save(business=self.request.user.business, created_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(business=self.request.user.business)


class SupplierPaymentViewSet(viewsets.ModelViewSet):
    serializer_class = SupplierPaymentSerializer
    permission_classes = [IsCashierOrAdmin]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['supplier', 'method']

    def get_queryset(self):
        return SupplierPayment.objects.select_related('supplier').filter(
            business=self.request.user.business
        )

    def perform_create(self, serializer):
        serializer.save(business=self.request.user.business, created_by=self.request.user)


class ExpenseReportView(APIView):
    permission_classes = [IsAdmin]

    def get(self, request):
        start = request.query_params.get('start_date', str(timezone.now().date().replace(day=1)))
        end = request.query_params.get('end_date', str(timezone.now().date()))
        expenses = Expense.objects.filter(
            expense_date__gte=start,
            expense_date__lte=end,
        ).select_related('category').order_by('-expense_date')
        summary = expenses.aggregate(
            total_amount=Sum('amount'),
            count=Count('id'),
            avg_amount=Avg('amount'),
        )
        by_category = list(
            expenses.values('category__name').annotate(
                total=Sum('amount'), count=Count('id')
            ).order_by('-total')
        )
        for row in by_category:
            row['category'] = row.pop('category__name') or 'Uncategorised'
        expense_list = list(expenses.values(
            'expense_date', 'description', 'amount', 'payment_method',
            'category__name', 'notes'
        ))
        for row in expense_list:
            row['category'] = row.pop('category__name') or ''
        report_format = request.query_params.get('export')
        if report_format in ['pdf', 'xlsx']:
            headers = ['Date', 'Description', 'Category', 'Method', 'Amount']
            rows = [
                [str(e['expense_date']), e['description'], e['category'], e['payment_method'], e['amount']]
                for e in expense_list
            ]
            summary_rows = [
                ('Total Expenses', summary['total_amount'] or 0),
                ('Number of Entries', summary['count'] or 0),
                ('Average per Entry', round(summary['avg_amount'] or 0, 2)),
            ]
            filename = f'expenses-report.{report_format}'
            if report_format == 'pdf':
                return _export_report_pdf('Expenses Report', headers, rows, filename, summary_rows=summary_rows, request_user=request.user)
            return _export_report_xlsx('Expenses Report', headers, rows, filename, summary_rows=summary_rows)
        return Response({
            'expenses': expense_list,
            'summary': {
                'total_amount': float(summary['total_amount'] or 0),
                'count': summary['count'] or 0,
                'avg_amount': round(float(summary['avg_amount'] or 0), 2),
            },
            'by_category': by_category,
        })


class SalesReturnViewSet(viewsets.ModelViewSet):
    serializer_class = SalesReturnSerializer
    permission_classes = [IsCashierOrAdmin]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['invoice', 'refund_method']
    search_fields = ['return_number', 'invoice__invoice_number', 'reason']
    ordering = ['-created_at']

    def get_queryset(self):
        return SalesReturn.objects.select_related('invoice', 'created_by').prefetch_related('items').filter(
            business=self.request.user.business
        )

    def create(self, request, *args, **kwargs):
        from django.db import transaction
        invoice_id = request.data.get('invoice')
        reason = request.data.get('reason', '')
        refund_method = request.data.get('refund_method', 'cash')
        items_data = request.data.get('items', [])  # [{invoice_item_id, quantity}]

        try:
            invoice = Invoice.objects.prefetch_related('items__product').get(
                pk=invoice_id, business=request.user.business
            )
        except Invoice.DoesNotExist:
            return Response({'detail': 'Invoice not found.'}, status=404)

        if invoice.status not in ('completed',):
            return Response({'detail': 'Only completed invoices can be returned.'}, status=400)

        if not items_data:
            return Response({'detail': 'Select at least one item to return.'}, status=400)

        with transaction.atomic():
            # Generate return number
            last = SalesReturn.objects.filter(business=request.user.business).order_by('-id').first()
            seq = (last.id + 1) if last else 1
            return_number = f"RET-{seq:04d}"

            refund_total = Decimal('0')
            return_obj = SalesReturn.objects.create(
                business=request.user.business,
                invoice=invoice,
                return_number=return_number,
                reason=reason,
                refund_method=refund_method,
                created_by=request.user,
            )

            for rd in items_data:
                try:
                    inv_item = InvoiceItem.objects.select_related('product').get(
                        pk=rd['invoice_item_id'], invoice=invoice
                    )
                except (InvoiceItem.DoesNotExist, KeyError):
                    return Response({'detail': f'Invalid item id {rd.get("invoice_item_id")}.'}, status=400)

                ret_qty = Decimal(str(rd.get('quantity', inv_item.quantity)))
                if ret_qty <= 0 or ret_qty > inv_item.quantity:
                    return Response({'detail': f'Invalid return qty for {inv_item.product_name}.'}, status=400)

                line_total = (inv_item.total / inv_item.quantity * ret_qty).quantize(Decimal('0.01'))
                refund_total += line_total

                SalesReturnItem.objects.create(
                    sales_return=return_obj,
                    invoice_item=inv_item,
                    product=inv_item.product,
                    product_name=inv_item.product_name,
                    quantity=ret_qty,
                    unit_price=inv_item.unit_price,
                    total=line_total,
                )

                # Restore stock
                if inv_item.product:
                    before = inv_item.product.current_stock
                    inv_item.product.current_stock += ret_qty
                    inv_item.product.save()
                    InventoryTransaction.objects.create(
                        business=request.user.business,
                        product=inv_item.product,
                        transaction_type='returned',
                        quantity=ret_qty,
                        before_stock=before,
                        after_stock=inv_item.product.current_stock,
                        reference=return_number,
                        created_by=request.user,
                    )

            return_obj.refund_amount = refund_total
            return_obj.save()

            # Mark invoice as refunded if full return
            total_returned = sum(
                Decimal(str(i.quantity)) for i in return_obj.items.all()
            )
            total_original = sum(
                Decimal(str(i.quantity)) for i in invoice.items.all()
            )
            if total_returned >= total_original:
                invoice.status = 'refunded'
                invoice.payment_status = 'refunded'
                invoice.save()

        return Response(SalesReturnSerializer(return_obj).data, status=201)


class BillsListView(APIView):
    """Alias: GET /bills/ returns invoices list (used by Returns page)."""
    permission_classes = [IsCashierOrAdmin]

    def get(self, request):
        qs = Invoice.objects.select_related('customer', 'created_by').prefetch_related('items', 'payments').all()
        user = request.user
        if user.role == 'cashier':
            qs = qs.filter(created_by=user)
        status_filter = request.query_params.get('status')
        payment_status = request.query_params.get('payment_status')
        if status_filter:
            qs = qs.filter(status=status_filter)
        if payment_status:
            qs = qs.filter(payment_status=payment_status)
        qs = qs.order_by('-created_at')[:200]
        return Response(InvoiceSerializer(qs, many=True).data)


class InvoicePDFView(APIView):
    permission_classes = []

    def get(self, request, pk):
        from .pdf_utils import generate_invoice_pdf, generate_thermal_invoice_pdf
        from rest_framework_simplejwt.tokens import AccessToken
        from django.contrib.auth import get_user_model

        # Allow token via query param for browser window.open() calls
        token = request.query_params.get('token')
        if token:
            try:
                validated = AccessToken(token)
                User = get_user_model()
                request.user = User.objects.get(id=validated['user_id'])
            except Exception:
                return HttpResponse('Invalid token', status=401)
        elif not request.user.is_authenticated:
            return HttpResponse('Unauthorized', status=401)

        try:
            invoice = Invoice.objects.prefetch_related('items', 'payments').select_related('customer').get(
                pk=pk, business=request.user.business
            )
        except Invoice.DoesNotExist:
            return HttpResponse('Not found', status=404)
        printer = request.query_params.get('printer')
        business_settings = {
            setting.key: setting.value
            for setting in Setting.objects.filter(business=invoice.business, key__in=['printer_type', 'invoice_template'])
        }
        printer_setting = business_settings.get('printer_type', 'a4').lower()
        template_setting = business_settings.get('invoice_template', 'gst_a4').lower()
        use_thermal = printer == 'thermal' or (
            not printer and (printer_setting in ('thermal', 'thermal_80', 'thermal_58') or template_setting.startswith('thermal'))
        )
        buffer = generate_thermal_invoice_pdf(invoice) if use_thermal else generate_invoice_pdf(invoice)
        response = HttpResponse(buffer, content_type='application/pdf')
        suffix = 'thermal' if use_thermal else 'a4'
        response['Content-Disposition'] = f'inline; filename="invoice-{invoice.invoice_number}-{suffix}.pdf"'
        return response


class CancelInvoiceView(APIView):
    permission_classes = [IsAdmin]

    def post(self, request, pk):
        from django.db import transaction
        try:
            invoice = Invoice.objects.get(pk=pk)
        except Invoice.DoesNotExist:
            return Response({'error': 'Not found'}, status=404)
        if invoice.status != 'completed':
            return Response({'error': 'Invoice already cancelled/refunded'}, status=400)
        with transaction.atomic():
            invoice.status = 'cancelled'
            invoice.save()
            # Restore stock
            for item in invoice.items.all():
                if item.product:
                    before = item.product.current_stock
                    item.product.current_stock += item.quantity
                    item.product.save()
                    InventoryTransaction.objects.create(
                        product=item.product,
                        transaction_type='returned',
                        quantity=item.quantity,
                        before_stock=before,
                        after_stock=item.product.current_stock,
                        reference=f"CANCEL-{invoice.invoice_number}",
                        created_by=request.user,
                    )
            # Reverse customer credit — only if this invoice contributed to outstanding
            if invoice.customer and invoice.payment_status in ('credit', 'partial') and invoice.balance_due > 0:
                invoice.customer.outstanding_amount = max(
                    Decimal('0'), invoice.customer.outstanding_amount - invoice.balance_due
                )
                invoice.customer.save()
        return Response({'status': 'cancelled'})


class RefundInvoiceView(APIView):
    permission_classes = [IsAdmin]

    def post(self, request, pk):
        try:
            invoice = Invoice.objects.get(pk=pk)
        except Invoice.DoesNotExist:
            return Response({'error': 'Not found'}, status=404)
        if invoice.status != 'completed':
            return Response({'error': 'Cannot refund'}, status=400)
        from django.db import transaction
        with transaction.atomic():
            invoice.status = 'refunded'
            invoice.save()
            for item in invoice.items.all():
                if item.product:
                    before = item.product.current_stock
                    item.product.current_stock += item.quantity
                    item.product.save()
                    InventoryTransaction.objects.create(
                        product=item.product,
                        transaction_type='returned',
                        quantity=item.quantity,
                        before_stock=before,
                        after_stock=item.product.current_stock,
                        reference=f"REFUND-{invoice.invoice_number}",
                        created_by=request.user,
                    )
        return Response({'status': 'refunded'})


class StockAdjustView(APIView):
    permission_classes = [IsAdmin]

    def post(self, request):
        product_id = request.data.get('product_id')
        quantity = Decimal(str(request.data.get('quantity', 0)))
        transaction_type = request.data.get('transaction_type', 'adjustment')
        notes = request.data.get('notes', '')
        try:
            product = Product.objects.get(pk=product_id)
        except Product.DoesNotExist:
            return Response({'error': 'Product not found'}, status=404)
        before = product.current_stock
        product.current_stock += quantity
        if product.current_stock < 0:
            product.current_stock = 0
        product.save()
        InventoryTransaction.objects.create(
            product=product,
            transaction_type=transaction_type,
            quantity=quantity,
            before_stock=before,
            after_stock=product.current_stock,
            notes=notes,
            created_by=request.user,
        )
        return Response(ProductSerializer(product).data)


class BulkStockAdjustView(APIView):
    """Apply several stock-in/out changes in one auditable transaction."""
    permission_classes = [IsAdmin]

    def post(self, request):
        from django.db import transaction
        items = request.data.get('items', [])
        if not isinstance(items, list) or not items:
            return Response({'detail': 'Add at least one product quantity.'}, status=400)
        with transaction.atomic():
            for item in items:
                try:
                    product = Product.objects.select_for_update().get(pk=item['product_id'])
                    quantity = Decimal(str(item['quantity']))
                except (Product.DoesNotExist, KeyError, ValueError, TypeError):
                    return Response({'detail': 'One or more stock rows are invalid.'}, status=400)
                if quantity == 0:
                    continue
                before = product.current_stock
                product.current_stock += quantity
                if product.current_stock < 0:
                    return Response({'detail': f'Insufficient stock for {product.name}.'}, status=400)
                product.save()
                InventoryTransaction.objects.create(
                    product=product,
                    transaction_type='stock_in' if quantity > 0 else 'stock_out',
                    quantity=quantity,
                    before_stock=before,
                    after_stock=product.current_stock,
                    reference='BULK-STOCK',
                    notes=request.data.get('notes', ''),
                    created_by=request.user,
                )
        return Response({'status': 'updated'})


class StockImportView(APIView):
    """Import stock changes from a CSV or modern Excel workbook."""
    permission_classes = [IsAdmin]

    MAX_IMPORT_ROWS = 1000

    def post(self, request):
        upload = request.FILES.get('file')
        if not upload:
            return Response({'detail': 'Choose a CSV or Excel file first.'}, status=400)
        if upload.size > 5 * 1024 * 1024:
            return Response({'detail': 'The import file must be smaller than 5 MB.'}, status=400)

        suffix = upload.name.rsplit('.', 1)[-1].lower() if '.' in upload.name else ''
        try:
            if suffix == 'csv':
                decoded = upload.read().decode('utf-8-sig')
                rows = list(csv.DictReader(decoded.splitlines()))
            elif suffix in ('xlsx', 'xlsm'):
                workbook = load_workbook(upload, read_only=True, data_only=True)
                sheet = workbook.active
                values = list(sheet.iter_rows(values_only=True))
                if not values:
                    rows = []
                else:
                    headers = [str(value).strip() if value is not None else '' for value in values[0]]
                    rows = [dict(zip(headers, values_row)) for values_row in values[1:] if any(value is not None and str(value).strip() for value in values_row)]
            else:
                return Response({'detail': 'Use a .csv, .xlsx, or .xlsm file.'}, status=400)
        except (UnicodeDecodeError, ValueError, OSError, TypeError) as exc:
            return Response({'detail': f'Could not read this file: {exc}'}, status=400)

        if not rows:
            return Response({'detail': 'The file has no stock rows.'}, status=400)
        if len(rows) > self.MAX_IMPORT_ROWS:
            return Response({'detail': f'Import up to {self.MAX_IMPORT_ROWS} rows at a time.'}, status=400)

        def normalise(row):
            return {str(key).strip().lower().replace(' ', '_').replace('-', '_'): value for key, value in row.items() if key is not None}

        prepared = []
        errors = []
        for row_number, raw_row in enumerate(rows, start=2):
            row = normalise(raw_row)
            sku = str(row.get('sku') or row.get('product_sku') or '').strip()
            product_name = str(row.get('product_name') or row.get('product') or row.get('name') or '').strip()
            raw_change = row.get('quantity', row.get('qty', row.get('stock_change', row.get('adjustment'))))
            raw_stock = row.get('current_stock', row.get('stock'))
            if not sku and not product_name:
                errors.append(f'Row {row_number}: enter a SKU or product name.')
                continue
            if raw_change in (None, '') and raw_stock in (None, ''):
                errors.append(f'Row {row_number}: enter Quantity to add/remove or Current Stock.')
                continue
            try:
                change = Decimal(str(raw_change)) if raw_change not in (None, '') else None
                stock = Decimal(str(raw_stock)) if raw_stock not in (None, '') else None
            except Exception:
                errors.append(f'Row {row_number}: stock values must be numbers.')
                continue
            if stock is not None and stock < 0:
                errors.append(f'Row {row_number}: Current Stock cannot be negative.')
                continue
            prepared.append((row_number, sku, product_name, change, stock))

        if errors:
            return Response({'detail': 'Fix the import file and try again.', 'errors': errors[:20]}, status=400)

        from django.db import transaction
        with transaction.atomic():
            resolved_rows = []
            for row_number, sku, product_name, change, stock in prepared:
                products = Product.objects.select_for_update()
                product = products.filter(business=request.user.business, sku__iexact=sku).first() if sku else products.filter(business=request.user.business, name__iexact=product_name).first()
                if not product:
                    lookup = f'SKU "{sku}"' if sku else f'product "{product_name}"'
                    return Response({'detail': f'Row {row_number}: no product found for {lookup}.'}, status=400)
                resolved_rows.append((row_number, product, change, stock))

            # Validate every resulting balance before changing stock, so an invalid
            # later row never leaves earlier rows partially imported.
            planned_rows = []
            running_stock = {}
            for row_number, product, change, stock in resolved_rows:
                before = running_stock.get(product.pk, product.current_stock)
                quantity = stock - before if stock is not None else change
                after = before + quantity
                if after < 0:
                    return Response({'detail': f'Row {row_number}: {product.name} would have negative stock.'}, status=400)
                running_stock[product.pk] = after
                planned_rows.append((product, quantity, before, after))

            imported = 0
            for product, quantity, before, after in planned_rows:
                if quantity == 0:
                    continue
                product.current_stock = after
                product.save(update_fields=['current_stock', 'updated_at'])
                InventoryTransaction.objects.create(
                    product=product,
                    transaction_type='stock_in' if quantity > 0 else 'stock_out',
                    quantity=quantity,
                    before_stock=before,
                    after_stock=after,
                    reference='STOCK-IMPORT',
                    notes=f'Imported from {upload.name}',
                    created_by=request.user,
                )
                imported += 1
        return Response({'status': 'updated', 'imported': imported, 'rows': len(prepared)})
